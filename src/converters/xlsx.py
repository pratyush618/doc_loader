from typing import Optional, BinaryIO
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime

from ..core.exceptions import ConversionException
from .base import BaseConverter, ConverterResult


class XLSXConverter(BaseConverter):
    """Converter for Microsoft Excel XLSX files (modern format)"""
    
    def accepts(self, file: BinaryIO, mimetype: Optional[str] = None,
                extension: Optional[str] = None) -> bool:
        """Check if this is an XLSX file"""
        if extension and extension.lower() in {'.xlsx', '.xlsm', '.xlsb'}:
            return True
        
        if mimetype and mimetype in {
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'application/vnd.ms-excel.sheet.macroEnabled.12',
            'application/vnd.ms-excel.sheet.binary.macroEnabled.12'
        }:
            return True
        
        # Try to open with openpyxl
        try:
            openpyxl.load_workbook(file, read_only=True)
            return True
        except (InvalidFileException, Exception):
            return False
        finally:
            self._reset_file_position(file)
    
    async def convert(self, file: BinaryIO, **kwargs) -> ConverterResult:
        """Convert XLSX to markdown"""
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            
            # Convert each sheet
            markdown_parts = []
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Add sheet header
                markdown_parts.append(f"# Sheet: {sheet_name}\n")
                
                # Check if sheet has data
                if worksheet.max_row == 1 and worksheet.max_column == 1:
                    # Check if the single cell has data
                    cell_value = worksheet.cell(1, 1).value
                    if cell_value is None:
                        markdown_parts.append("*Empty sheet*\n")
                        continue
                
                # Process sheet data
                sheet_md = await self._convert_sheet_to_markdown(worksheet)
                markdown_parts.append(sheet_md)
                markdown_parts.append("\n")
            
            content = "\n".join(markdown_parts)
            
            # Extract title from first sheet name or content
            title = None
            if len(workbook.sheetnames) == 1:
                title = f"Excel Sheet: {workbook.sheetnames[0]}"
            else:
                title = f"Excel Workbook ({len(workbook.sheetnames)} sheets)"
            
            return ConverterResult(
                content=content,
                title=title
            )
            
        except Exception as e:
            raise ConversionException(f"Failed to convert XLSX: {str(e)}")
    
    async def _convert_sheet_to_markdown(self, worksheet) -> str:
        """Convert a single worksheet to markdown table"""
        # Get the actual data range
        if worksheet.max_row == 1 and worksheet.max_column == 1:
            cell_value = worksheet.cell(1, 1).value
            if cell_value is None:
                return "*Empty sheet*"
        
        # Collect all rows with data
        rows = []
        max_col = 0
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                      min_col=1, max_col=worksheet.max_column):
            row_data = []
            has_data = False
            
            for cell in row:
                value = self._format_cell_value(cell.value)
                row_data.append(value)
                if value.strip():  # Non-empty cell
                    has_data = True
            
            # Only include rows that have at least one non-empty cell
            if has_data:
                rows.append(row_data)
                max_col = max(max_col, len(row_data))
        
        if not rows:
            return "*No data in sheet*"
        
        # Ensure all rows have the same number of columns
        for row in rows:
            while len(row) < max_col:
                row.append("")
        
        # Create markdown table
        markdown_lines = []
        
        # If we have at least one row, treat first row as header
        if rows:
            header = rows[0]
            
            # Create header row
            markdown_lines.append("| " + " | ".join(header) + " |")
            
            # Create separator row
            separators = []
            for cell in header:
                # Make separator at least 3 characters wide
                width = max(len(cell), 3)
                separators.append("-" * width)
            markdown_lines.append("| " + " | ".join(separators) + " |")
            
            # Add data rows
            for row in rows[1:]:
                markdown_lines.append("| " + " | ".join(row) + " |")
        
        return "\n".join(markdown_lines)
    
    def _format_cell_value(self, value) -> str:
        """Format cell value for markdown"""
        if value is None:
            return ""
        
        # Handle different data types
        if isinstance(value, bool):
            return str(value)
        elif isinstance(value, (int, float)):
            # Format numbers
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            return str(value)
        elif isinstance(value, datetime):
            # Format datetime
            return value.strftime("%Y-%m-%d %H:%M:%S")
        else:
            # Convert to string and escape pipe characters
            str_value = str(value)
            return str_value.replace("|", "\\|").replace("\n", " ").replace("\r", "")